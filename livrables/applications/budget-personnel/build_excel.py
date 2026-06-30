# -*- coding: utf-8 -*-
"""
Genere "Budget D-Jarvis.xlsx" : classeur de suivi budget propre, pro et simple,
fidele a la methodo de l'outil HTML (6 univers + 50/30/20), pre-rempli avec
l'historique (budget-data-2025-2026.json).

Saisie pour Damien = un seul menu deroulant "Categorie". Le classeur range tout
en coulisse (Type / Univers / Bucket / Mois) via formules.

Usage : py build_excel.py
"""
import json
from datetime import date
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, NamedStyle
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.formatting.rule import DataBarRule, CellIsRule
from openpyxl.chart import BarChart, LineChart, Reference

# --------------------------------------------------------------------------
# 0. Constantes & donnees
# --------------------------------------------------------------------------
HERE = "."
JSON_FILE = "budget-data-2025-2026.json"
OUT_FILE = "Budget D-Jarvis.xlsx"

# Palette inspiree des PDF de Damien
NAVY   = "1C2530"
ORANGE = "E8772E"
BLUE   = "2F6FB0"
GREEN  = "2FA37A"
GREY_L = "F2F4F7"
GREY_M = "E2E6EB"
RED    = "CF4646"
WHITE  = "FFFFFF"

EUR = '#,##0.00" "€'
PCT = '0.0%'

# Mapping canonique des categories (extrait de index.html)
# (label, type, univers, bucket, fixe)
CATS = [
    # --- Depenses perso : BESOINS ---
    ("logement",      "Logement & charges",            "Depense", "Perso",      "Besoins", True),
    ("alimentation",  "Courses & alimentation",        "Depense", "Perso",      "Besoins", False),
    ("transport",     "Transport",                     "Depense", "Perso",      "Besoins", False),
    ("factures",      "Factures fixes",                "Depense", "Perso",      "Besoins", True),
    ("sante",         "Sante",                         "Depense", "Perso",      "Besoins", False),
    # --- Depenses perso : ENVIES ---
    ("restaurants",   "Restaurants & sorties",         "Depense", "Perso",      "Envies",  False),
    ("shopping",      "Shopping & plaisir",            "Depense", "Perso",      "Envies",  False),
    ("loisirs",       "Voyages, loisirs & cadeaux",    "Depense", "Perso",      "Envies",  False),
    ("dons",          "Dons & solidarite",             "Depense", "Perso",      "Envies",  False),
    # --- Depenses perso : EPARGNE ---
    ("epargne",       "Epargne",                       "Depense", "Perso",      "Epargne", False),
    # --- Depenses hors train de vie ---
    ("avance",        "Avance pour un proche",         "Depense", "Avance",     "",        False),
    ("edrington",     "Frais Edrington (rembourses)",  "Depense", "Edrington",  "",        False),
    ("pro",           "Pro micro-entreprise",          "Depense", "Pro",        "",        False),
    ("patrimoine",    "Immobilier / parking",          "Depense", "Patrimoine", "",        False),
    ("bourse",        "Bourse & actions",              "Depense", "Patrimoine", "",        False),
    ("interne",       "Virement interne (exclu)",      "Depense", "Interne",    "",        False),
    # --- Revenus ---
    ("salaire",       "Salaire Edrington",             "Revenu",  "Reel",          "",     False),
    ("missions",      "Missions ponctuelles (Accor)",  "Revenu",  "Reel",          "",     False),
    ("micro",         "Revenus micro-entreprise",      "Revenu",  "Reel",          "",     False),
    ("aides",         "Aides (CAF, familial)",         "Revenu",  "Reel",          "",     False),
    ("remb_proche",   "Remboursement avance proche",   "Revenu",  "Remboursement", "",     False),
    ("remb_divers",   "Remboursement / retour divers", "Revenu",  "Remboursement", "",     False),
    ("remb_edrington","Remboursement frais Edrington", "Revenu",  "Remboursement", "",     False),
]
SLUG2LABEL = {c[0]: c[1] for c in CATS}

TARGETS = {"Besoins": 0.50, "Envies": 0.30, "Epargne": 0.20}
NET_PAR_CHAUFFEUR = 55.06

FR_MOIS = ["", "Janvier", "Fevrier", "Mars", "Avril", "Mai", "Juin",
           "Juillet", "Aout", "Septembre", "Octobre", "Novembre", "Decembre"]

# Mois de suivi : sept 2025 -> dec 2026 (colonnes futures pretes)
def month_range(y0, m0, y1, m1):
    out = []
    y, m = y0, m0
    while (y, m) <= (y1, m1):
        out.append((y, m))
        m += 1
        if m > 12:
            m = 1; y += 1
    return out
MONTHS = month_range(2025, 9, 2026, 12)            # 16 mois
MONTH_KEYS = [y * 100 + m for (y, m) in MONTHS]
MONTH_LABELS = [f"{FR_MOIS[m]} {y}" for (y, m) in MONTHS]

# --------------------------------------------------------------------------
# 1. Styles
# --------------------------------------------------------------------------
def fill(hexc): return PatternFill("solid", fgColor=hexc)
thin = Side(style="thin", color="D5DBE2")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

F_TITLE   = Font(name="Calibri", size=18, bold=True, color=NAVY)
F_SUB     = Font(name="Calibri", size=10, color="6B7682")
F_H       = Font(name="Calibri", size=11, bold=True, color=WHITE)
F_SECTION = Font(name="Calibri", size=11, bold=True, color=NAVY)
F_BOLD    = Font(name="Calibri", size=11, bold=True, color=NAVY)
F_KPI_LAB = Font(name="Calibri", size=9, bold=True, color="6B7682")
F_KPI_VAL = Font(name="Calibri", size=16, bold=True, color=NAVY)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT   = Alignment(horizontal="left", vertical="center")
RIGHT  = Alignment(horizontal="right", vertical="center")


def header_cell(ws, ref, text, fillc=NAVY):
    c = ws[ref]; c.value = text; c.font = F_H
    c.fill = fill(fillc); c.alignment = CENTER; c.border = BORDER
    return c


# --------------------------------------------------------------------------
# 2. Construction du classeur
# --------------------------------------------------------------------------
wb = Workbook()

# ===== Feuille PARAMETRES (placee tot pour les references) =================
wsP = wb.active
wsP.title = "Parametres"
wsP.sheet_view.showGridLines = False
wsP["A1"] = "Parametres"; wsP["A1"].font = F_TITLE
wsP["A2"] = "Tu peux ajuster ici : solde de depart, cibles, reperes. Les categories pilotent les menus deroulants."
wsP["A2"].font = F_SUB

# Solde de depart
wsP["A4"] = "Solde de depart (1er mois suivi)"; wsP["A4"].font = F_BOLD
wsP["B4"] = 1322.51; wsP["B4"].number_format = EUR; wsP["B4"].fill = fill(GREY_L)

# --- Bloc categories ---
CAT_HDR = 6
header_cell(wsP, f"A{CAT_HDR}", "Categorie (menu)")
header_cell(wsP, f"B{CAT_HDR}", "Type")
header_cell(wsP, f"C{CAT_HDR}", "Univers")
header_cell(wsP, f"D{CAT_HDR}", "Bucket 50/30/20")
header_cell(wsP, f"E{CAT_HDR}", "Charge fixe")
CAT_FIRST = CAT_HDR + 1
for i, (slug, label, typ, univ, bucket, fixe) in enumerate(CATS):
    r = CAT_FIRST + i
    wsP.cell(r, 1, label).border = BORDER
    wsP.cell(r, 2, typ).border = BORDER
    wsP.cell(r, 3, univ).border = BORDER
    wsP.cell(r, 4, bucket).border = BORDER
    wsP.cell(r, 5, "Oui" if fixe else "").border = BORDER
    band = GREY_L if i % 2 else WHITE
    for col in range(1, 6):
        wsP.cell(r, col).fill = fill(band)
CAT_LAST = CAT_FIRST + len(CATS) - 1

# --- Cibles 50/30/20 ---
T_HDR = CAT_LAST + 3
wsP.cell(T_HDR, 1, "Cibles 50/30/20").font = F_SECTION
wsP.cell(T_HDR + 1, 1, "Besoins"); wsP.cell(T_HDR + 1, 2, TARGETS["Besoins"]).number_format = PCT
wsP.cell(T_HDR + 2, 1, "Envies");  wsP.cell(T_HDR + 2, 2, TARGETS["Envies"]).number_format = PCT
wsP.cell(T_HDR + 3, 1, "Epargne"); wsP.cell(T_HDR + 3, 2, TARGETS["Epargne"]).number_format = PCT
for k in range(1, 4):
    wsP.cell(T_HDR + k, 2).fill = fill(GREY_L)
CIBLE_BESOINS = f"Parametres!$B${T_HDR+1}"
CIBLE_ENVIES  = f"Parametres!$B${T_HDR+2}"
CIBLE_EPARGNE = f"Parametres!$B${T_HDR+3}"

# --- Repere D-VTC ---
R_HDR = T_HDR + 5
wsP.cell(R_HDR, 1, "Repere D-VTC").font = F_SECTION
wsP.cell(R_HDR + 1, 1, "Net / mois par chauffeur")
wsP.cell(R_HDR + 1, 2, NET_PAR_CHAUFFEUR).number_format = EUR
wsP.cell(R_HDR + 1, 2).fill = fill(GREY_L)
NET_CHAUFFEUR = f"Parametres!$B${R_HDR+1}"

# --- Bloc mois ---
M_HDR = R_HDR + 3
header_cell(wsP, f"A{M_HDR}", "Mois (menu tableau de bord)")
header_cell(wsP, f"B{M_HDR}", "Cle")
M_FIRST = M_HDR + 1
for i, (lab, key) in enumerate(zip(MONTH_LABELS, MONTH_KEYS)):
    r = M_FIRST + i
    wsP.cell(r, 1, lab).border = BORDER
    wsP.cell(r, 2, key).border = BORDER
    band = GREY_L if i % 2 else WHITE
    wsP.cell(r, 1).fill = fill(band); wsP.cell(r, 2).fill = fill(band)
M_LAST = M_FIRST + len(MONTHS) - 1

wsP.column_dimensions["A"].width = 32
for col in "BCDE":
    wsP.column_dimensions[col].width = 16

# --- Noms definis ---
def add_name(name, ref):
    wb.defined_names.add(DefinedName(name, attr_text=ref))

add_name("soldeInitial",  "Parametres!$B$4")
add_name("catLabel",  f"Parametres!$A${CAT_FIRST}:$A${CAT_LAST}")
add_name("catType",   f"Parametres!$B${CAT_FIRST}:$B${CAT_LAST}")
add_name("catUniv",   f"Parametres!$C${CAT_FIRST}:$C${CAT_LAST}")
add_name("catBucket", f"Parametres!$D${CAT_FIRST}:$D${CAT_LAST}")
add_name("moisLabel", f"Parametres!$A${M_FIRST}:$A${M_LAST}")
add_name("moisKey",   f"Parametres!$B${M_FIRST}:$B${M_LAST}")

# ===== Feuille TRANSACTIONS (journal continu) ==============================
wsT = wb.create_sheet("Transactions")
wsT.sheet_view.showGridLines = False
wsT["A1"] = "Transactions"; wsT["A1"].font = F_TITLE
wsT["A2"] = "Saisis : Date, Montant, Description, puis choisis la Categorie. Le reste se remplit tout seul."
wsT["A2"].font = F_SUB

TH = 4   # ligne d'en-tete du tableau
cols = ["Date", "Montant", "Description", "Categorie", "Type", "Univers", "Bucket", "Mois"]
for j, name in enumerate(cols, start=1):
    header_cell(wsT, f"{get_column_letter(j)}{TH}", name)

# Donnees
data = json.load(open(JSON_FILE, encoding="utf-8"))["transactions"]
data.sort(key=lambda x: x["date"])
first = TH + 1
for i, tx in enumerate(data):
    r = first + i
    y, m, d = map(int, tx["date"].split("-"))
    cD = wsT.cell(r, 1, date(y, m, d)); cD.number_format = "DD/MM/YYYY"
    cM = wsT.cell(r, 2, round(float(tx["amount"]), 2)); cM.number_format = EUR
    wsT.cell(r, 3, tx.get("desc", ""))
    wsT.cell(r, 4, SLUG2LABEL.get(tx["cat"], tx["cat"]))
    # colonnes auto
    wsT.cell(r, 5, '=IFERROR(INDEX(catType,MATCH([@Categorie],catLabel,0)),"")')
    wsT.cell(r, 6, '=IFERROR(INDEX(catUniv,MATCH([@Categorie],catLabel,0)),"")')
    wsT.cell(r, 7, '=IFERROR(INDEX(catBucket,MATCH([@Categorie],catLabel,0)),"")')
    wsT.cell(r, 8, '=IF([@Date]="","",YEAR([@Date])*100+MONTH([@Date]))')
last = first + len(data) - 1

# Tableau Excel (auto-extension des formules sur les nouvelles lignes)
tbl = Table(displayName="Journal", ref=f"A{TH}:H{last}")
tbl.tableStyleInfo = TableStyleInfo(
    name="TableStyleLight9", showFirstColumn=False, showLastColumn=False,
    showRowStripes=True, showColumnStripes=False)
wsT.add_table(tbl)

# Menu deroulant Categorie (large marge pour les mois a venir)
dv = DataValidation(type="list", formula1="catLabel", allow_blank=True)
dv.error = "Choisis une categorie dans la liste."
dv.errorTitle = "Categorie invalide"
wsT.add_data_validation(dv)
dv.add(f"D{first}:D5000")

widths = {"A": 12, "B": 13, "C": 34, "D": 26, "E": 11, "F": 13, "G": 11, "H": 10}
for col, w in widths.items():
    wsT.column_dimensions[col].width = w
wsT.freeze_panes = f"A{first}"

JM = "Journal[Montant]"; JC = "Journal[Categorie]"; JMois = "Journal[Mois]"
JU = "Journal[Univers]"; JB = "Journal[Bucket]"

# ===== Feuille MOIS PAR MOIS (matrice) =====================================
wsM = wb.create_sheet("Mois par mois")
wsM.sheet_view.showGridLines = False
wsM["A1"] = "Mois par mois"; wsM["A1"].font = F_TITLE
wsM["A2"] = "Vue comparative. Tout se calcule automatiquement depuis l'onglet Transactions."
wsM["A2"].font = F_SUB

HDR_ROW = 4   # libelles
KEY_ROW = 5   # cles (servent aux calculs)
wsM.cell(HDR_ROW, 1, "Poste \\ Mois").font = F_H
wsM.cell(HDR_ROW, 1).fill = fill(NAVY); wsM.cell(HDR_ROW, 1).alignment = LEFT
wsM.cell(KEY_ROW, 1, "cle mois").font = F_SUB
ncol = len(MONTHS)
for i in range(ncol):
    col = get_column_letter(2 + i)
    hc = wsM[f"{col}{HDR_ROW}"]; hc.value = MONTH_LABELS[i]
    hc.font = F_H; hc.fill = fill(NAVY); hc.alignment = CENTER; hc.border = BORDER
    kc = wsM[f"{col}{KEY_ROW}"]; kc.value = MONTH_KEYS[i]
    kc.font = F_SUB; kc.alignment = CENTER
COLS = [get_column_letter(2 + i) for i in range(ncol)]

rows = {}      # nom logique -> numero de ligne
cur = KEY_ROW + 1

def section(title):
    global cur
    c = wsM.cell(cur, 1, title); c.font = F_H; c.fill = fill(ORANGE)
    c.alignment = LEFT
    for i in range(ncol):
        wsM.cell(cur, 2 + i).fill = fill(ORANGE)
    cur += 1

def cat_row(label, key=None, pct=False, bold=False, fillc=None):
    """Ligne dont chaque cellule est un SUMIFS par categorie+mois."""
    global cur
    rlabel = label
    lc = wsM.cell(cur, 1, rlabel)
    lc.font = F_BOLD if bold else Font(size=10, color=NAVY)
    lc.alignment = LEFT
    for i in range(ncol):
        col = COLS[i]
        cell = wsM.cell(cur, 2 + i)
        cell.value = (f'=SUMIFS({JM},{JC},$A{cur},{JMois},{col}${KEY_ROW})')
        cell.number_format = EUR
        if fillc: cell.fill = fill(fillc)
    rows[key or label] = cur
    cur += 1

def agg_row(label, key, field, crit, pct=False, bold=True, fillc=GREY_M):
    """Ligne SUMIFS par univers/bucket."""
    global cur
    lc = wsM.cell(cur, 1, label); lc.font = F_BOLD if bold else Font(size=10, color=NAVY)
    lc.alignment = LEFT; lc.fill = fill(fillc)
    for i in range(ncol):
        col = COLS[i]
        cell = wsM.cell(cur, 2 + i)
        cell.value = f'=SUMIFS({JM},{field},"{crit}",{JMois},{col}${KEY_ROW})'
        cell.number_format = EUR; cell.fill = fill(fillc)
    rows[key] = cur
    cur += 1

def formula_row(label, key, fn, number=EUR, bold=True, fillc=GREY_M):
    """Ligne calculee a partir d'autres lignes (fn(col)->formule)."""
    global cur
    lc = wsM.cell(cur, 1, label); lc.font = F_BOLD if bold else Font(size=10, color=NAVY)
    lc.alignment = LEFT; lc.fill = fill(fillc)
    for i in range(ncol):
        col = COLS[i]
        cell = wsM.cell(cur, 2 + i)
        cell.value = fn(col, i)
        cell.number_format = number; cell.fill = fill(fillc)
    rows[key] = cur
    cur += 1

# --- REVENUS ---
section("REVENUS")
for slug in ("salaire", "missions", "micro", "aides"):
    cat_row(SLUG2LABEL[slug], key=slug)
agg_row("Revenus reels (total)", "revReel", JU, "Reel", fillc=GREY_M)
agg_row("Remboursements recus", "remb", JU, "Remboursement", fillc=GREY_L)

# --- BESOINS ---
section("DEPENSES PERSO - BESOINS (cible 50%)")
for slug in ("logement", "alimentation", "transport", "factures", "sante"):
    cat_row(SLUG2LABEL[slug], key=slug)
agg_row("Besoins (total)", "besoins", JB, "Besoins", fillc=GREY_M)

# --- ENVIES ---
section("DEPENSES PERSO - ENVIES (cible 30%)")
for slug in ("restaurants", "shopping", "loisirs", "dons"):
    cat_row(SLUG2LABEL[slug], key=slug)
agg_row("Envies (total)", "envies", JB, "Envies", fillc=GREY_M)

# --- EPARGNE ---
section("EPARGNE (cible 20%)")
agg_row("Epargne (total)", "epargne", JB, "Epargne", fillc=GREY_M)

# --- HORS TRAIN DE VIE ---
section("HORS TRAIN DE VIE (neutralise)")
agg_row("Pro micro-entreprise", "pro", JU, "Pro", fillc=GREY_L)
agg_row("Frais Edrington (avances)", "edr", JU, "Edrington", fillc=GREY_L)
agg_row("Avances proches", "avance", JU, "Avance", fillc=GREY_L)
agg_row("Patrimoine (parking, bourse)", "patrimoine", JU, "Patrimoine", fillc=GREY_L)

# --- SYNTHESE ---
section("SYNTHESE DU MOIS")
formula_row("Cout de la vie (besoins+envies)", "coutVie",
            lambda col, i: f'={col}{rows["besoins"]}+{col}{rows["envies"]}', fillc=GREY_M)
formula_row("Depenses perso (+ epargne)", "depPerso",
            lambda col, i: f'={col}{rows["coutVie"]}+{col}{rows["epargne"]}', fillc=GREY_L)
formula_row("Reste / capacite d'epargne", "reste",
            lambda col, i: f'={col}{rows["revReel"]}-{col}{rows["coutVie"]}', fillc=GREY_M)
formula_row("% Besoins (/ revenu reel)", "pBes",
            lambda col, i: f'=IFERROR({col}{rows["besoins"]}/{col}{rows["revReel"]},0)',
            number=PCT, fillc=GREY_L)
formula_row("% Envies (/ revenu reel)", "pEnv",
            lambda col, i: f'=IFERROR({col}{rows["envies"]}/{col}{rows["revReel"]},0)',
            number=PCT, fillc=GREY_L)
formula_row("% Reste-epargne (/ revenu reel)", "pRes",
            lambda col, i: f'=IFERROR({col}{rows["reste"]}/{col}{rows["revReel"]},0)',
            number=PCT, fillc=GREY_L)
formula_row("Taux d'epargne reel", "tauxEp",
            lambda col, i: f'=IFERROR({col}{rows["epargne"]}/{col}{rows["revReel"]},0)',
            number=PCT, fillc=GREY_L)
formula_row("Flux du mois (tresorerie)", "flux",
            lambda col, i: (f'=({col}{rows["revReel"]}+{col}{rows["remb"]})'
                            f'-({col}{rows["depPerso"]}+{col}{rows["pro"]}'
                            f'+{col}{rows["edr"]}+{col}{rows["avance"]})'), fillc=GREY_M)

# Solde glissant
def solde_debut(col, i):
    if i == 0:
        return "=soldeInitial"
    prev = COLS[i - 1]
    return f'={prev}{rows["soldeFin"]}'
def solde_fin(col, i):
    return f'={col}{rows["soldeDebut"]}+{col}{rows["flux"]}'
# soldeFin reference soldeDebut ; on cree d'abord les 2 lignes puis on patch
formula_row("Solde DEBUT de mois", "soldeDebut", lambda col, i: "0", fillc=NAVY)
formula_row("Solde FIN de mois", "soldeFin", lambda col, i: "0", fillc=NAVY)
# patch contenu maintenant que les 2 numeros de ligne existent
for i in range(ncol):
    col = COLS[i]
    cd = wsM.cell(rows["soldeDebut"], 2 + i); cd.value = solde_debut(col, i)
    cd.font = Font(bold=True, color=WHITE); cd.number_format = EUR
    cf = wsM.cell(rows["soldeFin"], 2 + i); cf.value = solde_fin(col, i)
    cf.font = Font(bold=True, color=WHITE); cf.number_format = EUR
wsM.cell(rows["soldeDebut"], 1).font = Font(bold=True, color=WHITE)
wsM.cell(rows["soldeFin"], 1).font = Font(bold=True, color=WHITE)

wsM.column_dimensions["A"].width = 34
for c in COLS:
    wsM.column_dimensions[c].width = 13
wsM.freeze_panes = "B6"

# Graphiques (sous la matrice)
chart_anchor = cur + 2
# 1) Barres empilees Besoins / Envies / Epargne
bar = BarChart(); bar.type = "col"; bar.grouping = "stacked"; bar.overlap = 100
bar.title = "Repartition 50/30/20 par mois"
bar.height = 8; bar.width = 22
cats_ref = Reference(wsM, min_col=2, max_col=1 + ncol, min_row=HDR_ROW, max_row=HDR_ROW)
for key, name in (("besoins", "Besoins"), ("envies", "Envies"), ("epargne", "Epargne")):
    ref = Reference(wsM, min_col=1, max_col=1 + ncol, min_row=rows[key], max_row=rows[key])
    bar.add_data(ref, titles_from_data=True, from_rows=True)
bar.set_categories(cats_ref)
wsM.add_chart(bar, f"A{chart_anchor}")
# 2) Ligne Revenu reel vs Cout de la vie
line = LineChart(); line.title = "Revenu reel vs Cout de la vie"
line.height = 8; line.width = 22
for key in ("revReel", "coutVie"):
    ref = Reference(wsM, min_col=1, max_col=1 + ncol, min_row=rows[key], max_row=rows[key])
    line.add_data(ref, titles_from_data=True, from_rows=True)
line.set_categories(cats_ref)
wsM.add_chart(line, f"M{chart_anchor}")

# ===== Feuille TABLEAU DE BORD (mono-mois) =================================
wsD = wb.create_sheet("Tableau de bord")
wsD.sheet_view.showGridLines = False
wsD["B1"] = "Tableau de bord"; wsD["B1"].font = F_TITLE
wsD["B2"] = "Choisis un mois ci-dessous : tout se recalcule."; wsD["B2"].font = F_SUB

# Selecteur de mois
wsD["B4"] = "Mois :"; wsD["B4"].font = F_BOLD
sel = wsD["C4"]; sel.value = "Juin 2026"   # dernier mois avec donnees
sel.fill = fill(ORANGE); sel.font = Font(bold=True, color=WHITE, size=12)
sel.alignment = CENTER; sel.border = BORDER
dvm = DataValidation(type="list", formula1="moisLabel", allow_blank=False)
wsD.add_data_validation(dvm); dvm.add("C4")
# Cle du mois selectionne (cellule technique)
wsD["E4"] = "cle"; wsD["E4"].font = F_SUB
wsD["F4"] = "=INDEX(moisKey,MATCH($C$4,moisLabel,0))"
KEY = "$F$4"

def D_sumifs_univ(crit):
    return f'=SUMIFS({JM},{JU},"{crit}",{JMois},{KEY})'
def D_sumifs_bucket(crit):
    return f'=SUMIFS({JM},{JB},"{crit}",{JMois},{KEY})'

# Valeurs techniques (colonne H, masquee visuellement plus bas)
wsD["H1"] = "valeurs"; wsD["H1"].font = F_SUB
calc = {
    "revReel": D_sumifs_univ("Reel"),
    "remb":    D_sumifs_univ("Remboursement"),
    "besoins": D_sumifs_bucket("Besoins"),
    "envies":  D_sumifs_bucket("Envies"),
    "epargne": D_sumifs_bucket("Epargne"),
    "pro":     D_sumifs_univ("Pro"),
    "edr":     D_sumifs_univ("Edrington"),
    "avance":  D_sumifs_univ("Avance"),
    "micro":   f'=SUMIFS({JM},{JC},"{SLUG2LABEL["micro"]}",{JMois},{KEY})',
    "fixeLog": f'=SUMIFS({JM},{JC},"{SLUG2LABEL["logement"]}",{JMois},{KEY})',
    "fixeFac": f'=SUMIFS({JM},{JC},"{SLUG2LABEL["factures"]}",{JMois},{KEY})',
    "rembEdr": f'=SUMIFS({JM},{JC},"{SLUG2LABEL["remb_edrington"]}",{JMois},{KEY})',
    "rembPro": f'=SUMIFS({JM},{JC},"{SLUG2LABEL["remb_proche"]}",{JMois},{KEY})',
}
hrow = {}
r = 2
for k, f in calc.items():
    wsD.cell(r, 8, k).font = F_SUB
    wsD.cell(r, 9, f)
    hrow[k] = r; r += 1
H = lambda k: f"$I${hrow[k]}"
# derives
wsD.cell(r, 8, "coutVie"); wsD.cell(r, 9, f"={H('besoins')}+{H('envies')}"); hrow["coutVie"] = r; r += 1
wsD.cell(r, 8, "reste");   wsD.cell(r, 9, f"={H('revReel')}-{H('coutVie')}"); hrow["reste"] = r; r += 1
wsD.cell(r, 8, "fixe");    wsD.cell(r, 9, f"={H('fixeLog')}+{H('fixeFac')}"); hrow["fixe"] = r; r += 1
for k in hrow:
    wsD.cell(hrow[k], 9).number_format = EUR

# Solde debut/fin : INDEX dans la matrice
wsD.cell(r, 8, "soldeDebut")
wsD.cell(r, 9, f"=INDEX('Mois par mois'!$1:$1000,{rows['soldeDebut']},MATCH({KEY},'Mois par mois'!${KEY_ROW}:${KEY_ROW},0))")
hrow["soldeDebut"] = r; r += 1
wsD.cell(r, 8, "soldeFin")
wsD.cell(r, 9, f"=INDEX('Mois par mois'!$1:$1000,{rows['soldeFin']},MATCH({KEY},'Mois par mois'!${KEY_ROW}:${KEY_ROW},0))")
hrow["soldeFin"] = r; r += 1
wsD.cell(hrow["soldeDebut"], 9).number_format = EUR
wsD.cell(hrow["soldeFin"], 9).number_format = EUR

# --- KPI cards (ligne 6-8) ---
def kpi(col, lab, ref):
    c1 = wsD.cell(6, col); c1.value = lab; c1.font = F_KPI_LAB; c1.alignment = LEFT
    c1.fill = fill(GREY_L); c1.border = BORDER
    c2 = wsD.cell(7, col); c2.value = ref; c2.font = F_KPI_VAL; c2.alignment = LEFT
    c2.number_format = EUR; c2.fill = fill(GREY_L); c2.border = BORDER
    wsD.cell(8, col).fill = fill(GREY_L); wsD.cell(8, col).border = BORDER
kpi(2, "Tes revenus", f"={H('revReel')}")
kpi(3, "Pour vivre", f"={H('coutVie')}")
kpi(4, "Reste a epargner", f"={H('reste')}")
kpi(5, "Epargne placee", f"={H('epargne')}")
wsD.cell(8, 2, "= salaire + missions + micro + aides").font = F_SUB
wsD.cell(8, 3, "= besoins + envies").font = F_SUB
wsD.cell(8, 4, "= revenus - cout de la vie").font = F_SUB
wsD.cell(8, 5, "= depots livret / AV / tontine").font = F_SUB

# Solde glissant (ligne 10)
wsD.cell(10, 2, "Solde en debut de mois").font = F_BOLD
wsD.cell(10, 3, f"={H('soldeDebut')}").number_format = EUR
wsD.cell(10, 4, "Solde en fin de mois").font = F_BOLD
wsD.cell(10, 5, f"={H('soldeFin')}").number_format = EUR
wsD.cell(10, 5).font = Font(bold=True, color=NAVY)

# --- Bloc 50/30/20 (ligne 12+) ---
wsD.cell(12, 2, "Regle 50/30/20").font = F_SECTION
header_cell(wsD, "B13", "Poste"); header_cell(wsD, "C13", "Montant")
header_cell(wsD, "D13", "Part reelle"); header_cell(wsD, "E13", "Cible")
header_cell(wsD, "F13", "Ecart")
def rule_line(rr, name, hkey, cible_ref, direction):
    wsD.cell(rr, 2, name).font = F_BOLD
    wsD.cell(rr, 3, f"={H(hkey)}").number_format = EUR
    wsD.cell(rr, 4, f"=IFERROR({H(hkey)}/{H('revReel')},0)").number_format = PCT
    wsD.cell(rr, 5, f"={cible_ref}").number_format = PCT
    if direction == "max":   # besoins/envies : rester sous la cible
        wsD.cell(rr, 6, f"=E{rr}-D{rr}").number_format = PCT
    else:                    # reste : depasser la cible
        wsD.cell(rr, 6, f"=D{rr}-E{rr}").number_format = PCT
rule_line(14, "Besoins", "besoins", CIBLE_BESOINS, "max")
rule_line(15, "Envies", "envies", CIBLE_ENVIES, "max")
# reste-epargne : part reelle = reste/revenu
wsD.cell(16, 2, "Reste / epargne").font = F_BOLD
wsD.cell(16, 3, f"={H('reste')}").number_format = EUR
wsD.cell(16, 4, f"=IFERROR({H('reste')}/{H('revReel')},0)").number_format = PCT
wsD.cell(16, 5, f"={CIBLE_EPARGNE}").number_format = PCT
wsD.cell(16, 6, f"=D16-E16").number_format = PCT
# barres de donnees sur la part reelle
wsD.conditional_formatting.add("D14:D16",
    DataBarRule(start_type="num", start_value=0, end_type="num", end_value=1,
                color=BLUE, showValue=True))
# ecart negatif en rouge
wsD.conditional_formatting.add("F14:F16",
    CellIsRule(operator="lessThan", formula=["0"], font=Font(color=RED, bold=True)))

# --- Argent qui passe (ligne 18+) ---
wsD.cell(18, 2, "Argent qui ne fait que passer").font = F_SECTION
wsD.cell(19, 2, "Frais Edrington avances ce mois").font = F_BOLD
wsD.cell(19, 3, f"={H('edr')}").number_format = EUR
wsD.cell(20, 2, "Remboursement Edrington recu").font = F_BOLD
wsD.cell(20, 3, f"={H('rembEdr')}").number_format = EUR
wsD.cell(21, 2, "Avances a des proches ce mois").font = F_BOLD
wsD.cell(21, 3, f"={H('avance')}").number_format = EUR
wsD.cell(22, 2, "Remboursements proches recus").font = F_BOLD
wsD.cell(22, 3, f"={H('rembPro')}").number_format = EUR

# --- Objectif charges fixes (ligne 24+) ---
wsD.cell(24, 2, "Objectif : couvrir les charges fixes avec la micro").font = F_SECTION
wsD.cell(25, 2, "Charges fixes du mois").font = F_BOLD
wsD.cell(25, 3, f"={H('fixe')}").number_format = EUR
wsD.cell(26, 2, "Couvert par revenus micro").font = F_BOLD
wsD.cell(26, 3, f"={H('micro')}").number_format = EUR
wsD.cell(27, 2, "Reste a couvrir").font = F_BOLD
wsD.cell(27, 3, f"=MAX(0,{H('fixe')}-{H('micro')})").number_format = EUR
wsD.cell(28, 2, "= chauffeurs D-VTC equivalents").font = F_BOLD
wsD.cell(28, 3, f"=ROUNDUP(MAX(0,{H('fixe')}-{H('micro')})/{NET_CHAUFFEUR},0)")
wsD.cell(28, 3).font = Font(bold=True, color=ORANGE, size=14)

# largeurs + masquage colonnes techniques
for col, w in {"A": 3, "B": 32, "C": 16, "D": 13, "E": 13, "F": 13}.items():
    wsD.column_dimensions[col].width = w
wsD.column_dimensions["H"].hidden = True
wsD.column_dimensions["I"].hidden = True

# ===== Feuille LISEZ-MOI ===================================================
wsR = wb.create_sheet("Lisez-moi")
wsR.sheet_view.showGridLines = False
wsR["B2"] = "Budget D-Jarvis - mode d'emploi"; wsR["B2"].font = F_TITLE
guide = [
    "",
    "1. Onglet TRANSACTIONS : ta seule zone de saisie.",
    "   Tape Date, Montant, Description, puis choisis la Categorie dans le menu deroulant.",
    "   Les colonnes Type / Univers / Bucket / Mois se remplissent toutes seules.",
    "",
    "2. Onglet TABLEAU DE BORD : choisis un mois en haut, tout se recalcule",
    "   (KPI, regle 50/30/20, argent qui passe, objectif chauffeurs D-VTC).",
    "",
    "3. Onglet MOIS PAR MOIS : la vue comparative + graphiques d'evolution.",
    "",
    "4. Onglet PARAMETRES : solde de depart, cibles 50/30/20, reperes, liste des categories.",
    "",
    "Bon a savoir :",
    "- Les frais Pro / Edrington avances et les avances a des proches sont neutralises :",
    "  ils ne polluent pas ton train de vie (50/30/20).",
    "- Les virements internes (Virement interne) et le patrimoine (parking, bourse) sont",
    "  exclus du calcul 50/30/20 pour ne pas fausser tes chiffres.",
    "- Le solde de chaque mois part du solde de fin du mois precedent (tu fixes seulement",
    "  le tout premier dans Parametres).",
]
for i, line in enumerate(guide):
    c = wsR.cell(4 + i, 2, line)
    c.font = F_BOLD if line and not line.startswith(" ") and line[0].isdigit() else Font(size=11, color=NAVY)
wsR.column_dimensions["A"].width = 3
wsR.column_dimensions["B"].width = 95

# Ordre des onglets
wb.move_sheet("Lisez-moi", -(len(wb.sheetnames) - 1))   # en premier
wb._sheets.sort(key=lambda s: ["Lisez-moi", "Tableau de bord", "Mois par mois",
                               "Transactions", "Parametres"].index(s.title))
wb.active = wb.sheetnames.index("Tableau de bord")

wb.save(OUT_FILE)
print(f"OK -> {OUT_FILE}  |  {len(data)} transactions  |  {len(MONTHS)} mois")
